""" excelhelper.py
	Patryk Sienkiewicz (Patsen95), 2022
	https://github.com/Patsen95

	This file contains wrapping functions for OpenPyXL modules and adds some new features.
"""

from openpyxl.styles import NamedStyle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from copy import copy



def getCellCoord(cellCoord: str) -> int:
	"""Convert Excel's alphanumeric address to column / row indexes.

	Args:
		cellCoord: str - alphanumeric address

	Returns:
		Tuple with column and row indexes
	"""
	_coord = coordinate_from_string(cellCoord)
	_col = column_index_from_string(_coord[0])
	_row = _coord[1]
	return _col, _row


def getXlAddress(column: int, row: int, zeroIndexed=False) -> str:
	"""Convert column / row index to Excel's alphanumeric address.

	Args:
		column: int - cell's column
		row: int - cell's row
		zeroIndexed: bool - (optional) if True, row & column indexes count from zero

	Returns:
		Alphanumeric address as string
	"""
	if zeroIndexed:
		row += 1
		column += 1
	return get_column_letter(column) + str(row)


def copyStyle(sheet: Workbook, fromCell: str, toCell: str = None): #-> NamedStyle:
	"""Copy cell's formatting from selected range.
	Cells are addressed by strings like 'A3'.

	Args:
		sheet: Workbook - selected sheet
		fromCell: str - lower cell address
		toCell: str - (optional) higher cell address

	Returns:
		Tuple as container whose first two indexes are addresses (also as tuple) start and end
		of selected cells range, and a third index is a list of NamedStyle objects
	"""
	_start = getCellCoord(fromCell)
	_end = 0

	if toCell == None: _end = _start
	else: _end = getCellCoord(toCell)

	_styles = []

	for row in sheet.iter_rows(min_col=_start[0], min_row=_start[1], max_col=_end[0], max_row=_end[1]):
		for cell in row:
			_ns = NamedStyle()
			_ns.font = copy(cell.font)
			_ns.border = copy(cell.border)
			_ns.fill = copy(cell.fill)
			_ns.number_format = copy(cell.number_format)
			_ns.protection = copy(cell.protection)
			_ns.alignment = copy(cell.alignment)
				
			_styles.append(_ns)
	return (_start, _end, _styles)


def applyStyle(sheet: Workbook, style: tuple, fromCell: str=None, toCell: str=None):
	"""Apply formatting and style to selected cells. Style can be applied
	automatically to the same cells from which has been copied or to other selected cells
	using optional parameters.

	Args:
		sheet: Workbook - selected sheet
		style: tuple - container with cells range and NameStyle object list
		fromCell: str - (optional) lower cell address
		toCell: str - (optional) higher cell address

	Returns:
		None
	"""
	if (fromCell == None) or (toCell == None):
		_start = style[0]
		_end = style[1]
	else:
		_start = getCellCoord(fromCell)
		_end = getCellCoord(toCell)

	_styles = style[2]
	_cellIdx = 0

	for row in sheet.iter_rows(min_col=_start[0], min_row=_start[1], max_col=_end[0], max_row=_end[1]):
		_style = _styles[_cellIdx]
		for cell in row:
			cell.font = copy(_style.font)
			cell.border = copy(_style.border)
			cell.fill = copy(_style.fill)
			cell.number_format = copy(_style.number_format)
			cell.protection = copy(_style.protection)
			cell.alignment = copy(_style.alignment)
		_cellIdx += 1
			
